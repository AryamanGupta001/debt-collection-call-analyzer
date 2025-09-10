# run_batch.py
import argparse
import csv
import json
from pathlib import Path
from src.io_json import load_file
from src.profanity import detect_profanity
from src.pii_compliance import detect_compliance_violation
from src.metrics import overtalk_percentage, silence_percentage
from src.metrics import talk_share

# Excel support (optional)
try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Alignment
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

def process_file(path: Path, strict=False):
    """Process a single transcript file and return analysis results."""
    try:
        utt = load_file(path)
    except Exception as e:
        return {"call_id": path.stem, "error": str(e)}

    # Calculate metrics
    ot = overtalk_percentage(utt)
    si = silence_percentage(utt)
    tt = talk_share(utt)

    # Run detection
    prof = detect_profanity(utt)
    comp = detect_compliance_violation(utt, strict=strict)

    return {
        "call_id": Path(path).stem,
        "agent_prof": "Yes" if prof.get("agent_has") else "No",
        "borrower_prof": "Yes" if prof.get("borrower_has") else "No",
        "compliance_violation": "Yes" if comp.get("violation") else "No",
        "overtalk_pct": f"{ot:.2f}%",
        "silence_pct": f"{si:.2f}%",
        "total_time": f"{tt['total']:.1f}s",
        "agent_share": f"{tt['agent_pct']:.1f}%",
        "borrower_share": f"{tt['borrower_pct']:.1f}%",
        "prof_details": prof,
        "comp_details": comp,
        "raw_metrics": {
            "overtalk": ot,
            "silence": si
        }
    }

def create_formatted_excel(csv_path, excel_path):
    """Create a formatted Excel file from a CSV file."""
    if not EXCEL_SUPPORT:
        return False
        
    try:
        # Read CSV into pandas
        df = pd.read_csv(csv_path)
        
        # Write to Excel
        df.to_excel(excel_path, index=False)
        
        # Load workbook to format it
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        # Set column widths
        col_widths = {
            "File (Call ID)": 40,
            "call_id": 40,
            "Agent Profanity": 15,
            "Borrower Profanity": 15,
            "flag": 10,
            "overtalk_pct": 15,
            "silence_pct": 15,
            "Overtalk %": 15,
            "Silence %": 15,
            "details": 50
        }
        
        for i, col_name in enumerate(df.columns):
            col_letter = openpyxl.utils.get_column_letter(i+1)
            width = col_widths.get(col_name, 20)
            ws.column_dimensions[col_letter].width = width
            
        # Enable text wrapping
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
                
        # Save the formatted workbook
        wb.save(excel_path)
        return True
    except Exception as e:
        print(f"⚠️ Excel formatting error: {str(e)}")
        return False

def main():
    ap = argparse.ArgumentParser(description="Batch process call transcripts")
    ap.add_argument("--input_dir", required=True, help="Directory containing JSON/YAML files")
    ap.add_argument("--strict", action="store_true", help="Enable strict compliance verification")
    ap.add_argument("--profanity", action="store_true", help="Only check for profanity (skip compliance checks)")
    ap.add_argument("--no_excel", action="store_true", help="Skip Excel file generation")
    args = ap.parse_args()

    input_path = Path(args.input_dir)
    results_dir = Path("results")
    results_dir.mkdir(exist_ok=True)

    files = list(input_path.glob("**/*.json")) + list(input_path.glob("**/*.yaml")) + list(input_path.glob("**/*.yml"))
    
    if not files:
        print(f"⚠️ No JSON/YAML files found in {input_path}")
        return

    # Process all files
    if args.profanity:
        # PROFANITY MODE - Special format for profanity detection
        summary_rows = []
        detail_rows = []  # Add this line to collect detail rows

        for f in files:
            try:
                # Load file and run profanity detection directly
                utt = load_file(f)
                prof = detect_profanity(utt)
                
                # Calculate metrics
                ot = overtalk_percentage(utt)
                si = silence_percentage(utt)
                
                # Format as seen in results_profanity.csv
                flag = "Yes" if (prof.get("agent_has", False) or prof.get("borrower_has", False)) else "No"
                summary_rows.append({
                    "call_id": f.stem,
                    "flag": flag,
                    "overtalk_pct": ot,
                    "silence_pct": si,
                    "details": json.dumps(prof)
                })
                
                # Add details for each profanity hit
                for hit in prof.get("hits", []):
                    detail_rows.append({
                        "call_id": f.stem,
                        "speaker": hit.get("speaker", ""),
                        "text": hit.get("text", ""),
                        "time": f"{hit.get('stime', 0):.1f}s",
                        "matches": ", ".join(hit.get("matches", []))
                    })
            except Exception as e:
                print(f"⚠️ Skipping {f.name}: {str(e)}")
        
        # Write profanity CSV
        if summary_rows:
            summary_file = results_dir / "summary_profanity.csv"
            with open(summary_file, "w", newline="", encoding="utf-8") as fh:
                writer = csv.DictWriter(fh, ["call_id", "flag", "overtalk_pct", "silence_pct", "details"])
                writer.writeheader()
                writer.writerows(summary_rows)
                
            print(f"✅ Profanity summary saved to {summary_file}")
            
            # Create Excel if requested
            if EXCEL_SUPPORT and not args.no_excel:
                excel_summary_file = str(summary_file).replace('.csv', '.xlsx')
                if create_formatted_excel(summary_file, excel_summary_file):
                    print(f"✅ Formatted Excel saved to {Path(excel_summary_file).name}")

        # Write details CSV for profanity hits
        if detail_rows:
            details_file = results_dir / "details_profanity.csv"
            with open(details_file, "w", newline="", encoding="utf-8") as fh:
                writer = csv.DictWriter(fh, ["call_id", "speaker", "text", "time", "matches"])
                writer.writeheader()
                writer.writerows(detail_rows)
                
            print(f"✅ Profanity details saved to {details_file}")
            
            # Create Excel if requested
            if EXCEL_SUPPORT and not args.no_excel:
                excel_details_file = str(details_file).replace('.csv', '.xlsx')
                if create_formatted_excel(details_file, excel_details_file):
                    print(f"✅ Formatted Excel saved to {Path(excel_details_file).name}")
    else:
        # STANDARD MODE - Full analysis
        summary_rows = []
        detail_rows = []
        
        for f in files:
            # Process each file
            res = process_file(f, strict=args.strict)
            
            if "error" in res:
                print(f"⚠️ Skipping {f.name}: {res['error']}")
                continue
            
            # Add to summary rows
            summary_rows.append({
                "File (Call ID)": res["call_id"],
                "Agent Profanity": res["agent_prof"],
                "Borrower Profanity": res["borrower_prof"],
                "Compliance Violation": res["compliance_violation"],
                "Overtalk %": res["overtalk_pct"],
                "Silence %": res["silence_pct"],
                "Total Time": res["total_time"],
                "Agent Talk %": res["agent_share"],
                "Borrower Talk %": res["borrower_share"]
            })

            # Add details for profanity
            for hit in res["prof_details"].get("hits", []):
                detail_rows.append({
                    "File (Call ID)": res["call_id"],
                    "Type": "Profanity",
                    "Speaker": hit.get("speaker", ""),
                    "Text": hit.get("text", ""),
                    "Matches": ", ".join(hit.get("matches", []))
                })

            # Add details for compliance violations
            ev = res["comp_details"].get("evidence", {})
            if isinstance(ev, dict) and ev.get("reason"):
                detail_rows.append({
                    "File (Call ID)": res["call_id"],
                    "Type": "Compliance",
                    "Speaker": "agent",
                    "Text": ev.get("reason", ""),
                    "Matches": json.dumps(ev.get("examples", []))
                })

        # Determine output filenames based on mode
        summary_file = results_dir / ("summary_strict.csv" if args.strict else "summary.csv")
        details_file = results_dir / ("details_strict.csv" if args.strict else "details.csv")

        # Write summary CSV
        try:
            if summary_rows:
                with open(summary_file, "w", newline="", encoding="utf-8") as fh:
                    writer = csv.DictWriter(fh, summary_rows[0].keys())
                    writer.writeheader()
                    writer.writerows(summary_rows)
                    
                print(f"✅ Summary saved to {summary_file}")
                
                # Create Excel if requested
                if EXCEL_SUPPORT and not args.no_excel:
                    excel_summary_file = str(summary_file).replace('.csv', '.xlsx')
                    if create_formatted_excel(summary_file, excel_summary_file):
                        print(f"✅ Formatted Excel saved to {Path(excel_summary_file).name}")
            else:
                print("⚠️ No valid results to save in summary")
        except Exception as e:
            print(f"❌ Error writing summary CSV: {str(e)}")

        # Write details CSV
        try:
            if detail_rows:
                with open(details_file, "w", newline="", encoding="utf-8") as fh:
                    writer = csv.DictWriter(fh, detail_rows[0].keys())
                    writer.writeheader()
                    writer.writerows(detail_rows)
                    
                print(f"✅ Details saved to {details_file}")
                
                # Create Excel if requested
                if EXCEL_SUPPORT and not args.no_excel:
                    excel_details_file = str(details_file).replace('.csv', '.xlsx')
                    if create_formatted_excel(details_file, excel_details_file):
                        print(f"✅ Formatted Excel saved to {Path(excel_details_file).name}")
            else:
                print("ℹ️ No detail rows to save")
        except Exception as e:
            print(f"❌ Error writing details CSV: {str(e)}")

    # Summary statistics
    print(f"📊 Processed {len(files)} files")

if __name__ == "__main__":
    main()
